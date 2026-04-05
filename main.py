#!/usr/bin/env python3
"""확률통계 채점 시스템 — Excel 가로 입력 → SQLite 세로 저장 → LMS/피드백 출력"""

import argparse
import re
import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DB_PATH = Path("db/grades.db")
INPUT_EXCEL = Path("input/grades.xlsx")

# ── styling constants ───────────────────────────────

HEADER_FONT = Font(name="Arial", bold=True, size=11)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
SCORE_FILL = PatternFill("solid", fgColor="E2EFDA")
DEDUCT_FILL = PatternFill("solid", fgColor="FCE4D6")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center")
WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ── DB ──────────────────────────────────────────────

DDL = """
CREATE TABLE IF NOT EXISTS students (
    학번 INTEGER PRIMARY KEY,
    이름 TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS problems (
    평가유형 TEXT NOT NULL,
    회차     INTEGER NOT NULL,
    문제번호 INTEGER NOT NULL,
    만점     INTEGER NOT NULL,
    정답_소수 TEXT,
    정답_분수 TEXT,
    채점기준 TEXT,
    PRIMARY KEY (평가유형, 회차, 문제번호)
);

CREATE TABLE IF NOT EXISTS submissions (
    학번     INTEGER NOT NULL,
    평가유형 TEXT NOT NULL,
    회차     INTEGER NOT NULL,
    제출여부 TEXT NOT NULL CHECK (제출여부 IN ('O', 'X')),
    제출시각 TEXT,
    지각유형 TEXT NOT NULL CHECK (지각유형 IN ('정상', '1형', '2형')),
    PRIMARY KEY (학번, 평가유형, 회차),
    FOREIGN KEY (학번) REFERENCES students(학번)
);

CREATE TABLE IF NOT EXISTS grades (
    학번     INTEGER NOT NULL,
    평가유형 TEXT NOT NULL,
    회차     INTEGER NOT NULL,
    문제번호 INTEGER NOT NULL,
    점수     REAL,
    감점이유 TEXT,
    PRIMARY KEY (학번, 평가유형, 회차, 문제번호),
    FOREIGN KEY (학번) REFERENCES students(학번),
    FOREIGN KEY (평가유형, 회차, 문제번호)
        REFERENCES problems(평가유형, 회차, 문제번호)
);
"""


def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def _styled_cell(ws, row, col, value, fill=None, font=None, align=CENTER):
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = THIN_BORDER
    cell.alignment = align
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    return cell


# ── COMMANDS ────────────────────────────────────────

def cmd_init(_args):
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = get_conn()
    conn.executescript(DDL)
    conn.close()
    print(f"[OK] DB 초기화 완료: {DB_PATH}")


def cmd_import_students(args):
    df = pd.read_excel(args.file, sheet_name="수강생", dtype={"학번": int})
    conn = get_conn()
    conn.execute("BEGIN")
    for _, row in df.iterrows():
        conn.execute(
            "INSERT OR REPLACE INTO students (학번, 이름) VALUES (?, ?)",
            (int(row["학번"]), str(row["이름"])),
        )
    conn.commit()
    conn.close()
    print(f"[OK] 수강생 {len(df)}명 등록 완료")


def cmd_import_problems(args):
    df = pd.read_excel(args.file, sheet_name="문제정보")
    conn = get_conn()
    conn.execute("BEGIN")
    for _, row in df.iterrows():
        conn.execute(
            """INSERT OR REPLACE INTO problems
               (평가유형, 회차, 문제번호, 만점, 정답_소수, 정답_분수, 채점기준)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (
                str(row["평가유형"]),
                int(row["회차"]),
                int(row["문제번호"]),
                int(row["만점"]),
                None if pd.isna(row.get("정답_소수")) else str(row["정답_소수"]),
                None if pd.isna(row.get("정답_분수")) else str(row["정답_분수"]),
                None if pd.isna(row.get("채점기준")) else str(row["채점기준"]),
            ),
        )
    conn.commit()
    conn.close()
    print(f"[OK] 문제정보 {len(df)}건 등록 완료")


def cmd_import_submissions(args):
    df = pd.read_excel(args.file, sheet_name="제출현황")
    conn = get_conn()
    conn.execute("BEGIN")
    for _, row in df.iterrows():
        conn.execute(
            """INSERT OR REPLACE INTO submissions
               (학번, 평가유형, 회차, 제출여부, 제출시각, 지각유형)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (
                int(row["학번"]),
                str(row["평가유형"]),
                int(row["회차"]),
                str(row["제출여부"]),
                None if pd.isna(row.get("제출시각")) else str(row["제출시각"]),
                str(row["지각유형"]),
            ),
        )
    conn.commit()
    conn.close()
    print(f"[OK] 제출현황 {len(df)}건 등록 완료")


def cmd_gen_template(args):
    eval_type, round_num = args.type, args.round
    conn = get_conn()

    problems = pd.read_sql(
        "SELECT * FROM problems WHERE 평가유형=? AND 회차=? ORDER BY 문제번호",
        conn,
        params=[eval_type, round_num],
    )
    if problems.empty:
        print(f"[ERROR] {eval_type} {round_num}회차 문제정보가 없습니다.")
        conn.close()
        return

    students = pd.read_sql("SELECT 학번, 이름 FROM students ORDER BY 학번", conn)
    if students.empty:
        print("[ERROR] 수강생 명단이 없습니다.")
        conn.close()
        return

    n_prob = len(problems)

    # 터미널에 채점 기준 출력
    print(f"\n{'='*60}")
    print(f"  {eval_type} {round_num}회차 — 문제 {n_prob}개")
    print(f"{'='*60}")
    for _, p in problems.iterrows():
        ans_d = p["정답_소수"] if pd.notna(p["정답_소수"]) else "-"
        ans_f = p["정답_분수"] if pd.notna(p["정답_분수"]) else "-"
        crit = p["채점기준"] if pd.notna(p["채점기준"]) else "-"
        print(
            f"  문제{int(p['문제번호']):>2} | 만점: {int(p['만점']):>3}"
            f" | 소수: {ans_d} | 분수: {ans_f} | 기준: {crit}"
        )
    print(f"{'='*60}\n")

    # ── Sheet 1: 채점 (가로 템플릿) ──
    wb = Workbook()
    ws = wb.active
    ws.title = "채점"

    headers = ["학번", "이름"]
    for p_num in range(1, n_prob + 1):
        headers.extend([f"{p_num}_점수", f"{p_num}_감점"])

    for col, h in enumerate(headers, 1):
        fill = HEADER_FILL
        if "_점수" in str(h):
            fill = SCORE_FILL
        elif "_감점" in str(h):
            fill = DEDUCT_FILL
        _styled_cell(ws, 1, col, h, fill=fill, font=HEADER_FONT)

    for r, (_, stu) in enumerate(students.iterrows(), 2):
        _styled_cell(ws, r, 1, int(stu["학번"]))
        _styled_cell(ws, r, 2, stu["이름"])
        for c in range(3, len(headers) + 1):
            _styled_cell(ws, r, c, None)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 10
    for c in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12
    ws.freeze_panes = "C2"

    # ── Sheet 2: 채점기준(참고) ──
    ws2 = wb.create_sheet("채점기준(참고)")
    ref_headers = ["문제번호", "만점", "정답(소수)", "정답(분수)", "채점기준"]
    for col, h in enumerate(ref_headers, 1):
        _styled_cell(ws2, 1, col, h, fill=HEADER_FILL, font=HEADER_FONT)

    for r, (_, p) in enumerate(problems.iterrows(), 2):
        _styled_cell(ws2, r, 1, int(p["문제번호"]))
        _styled_cell(ws2, r, 2, int(p["만점"]))
        _styled_cell(ws2, r, 3, p["정답_소수"] if pd.notna(p["정답_소수"]) else "")
        _styled_cell(ws2, r, 4, p["정답_분수"] if pd.notna(p["정답_분수"]) else "")
        _styled_cell(
            ws2, r, 5,
            p["채점기준"] if pd.notna(p["채점기준"]) else "",
            align=WRAP,
        )

    ws2.column_dimensions["A"].width = 10
    ws2.column_dimensions["B"].width = 8
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 14
    ws2.column_dimensions["E"].width = 30
    ws2.protection.sheet = True

    out_path = Path(f"rounds/{eval_type}_{round_num}/template_{eval_type}_{round_num}.xlsx")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    conn.close()
    print(f"[OK] 채점 템플릿 생성: {out_path}")
    print(f"     채점 시트: 학생 {len(students)}명 × 문제 {n_prob}개")
    print(f"     채점기준(참고) 시트: 정답 및 채점기준 포함")


def cmd_import_grades(args):
    template_path = Path(args.file)

    match = re.match(r"template_(.+)_(\d+)\.xlsx", template_path.name)
    if match:
        eval_type = match.group(1)
        round_num = int(match.group(2))
    elif args.type and args.round:
        eval_type, round_num = args.type, args.round
    else:
        print("[ERROR] 파일명에서 평가유형/회차를 추출할 수 없습니다.")
        print("  파일명 형식: template_<평가유형>_<회차>.xlsx")
        print("  또는 --type, --round 옵션을 지정하세요.")
        return

    df = pd.read_excel(template_path, sheet_name="채점")
    conn = get_conn()

    problems = pd.read_sql(
        "SELECT 문제번호 FROM problems WHERE 평가유형=? AND 회차=? ORDER BY 문제번호",
        conn,
        params=[eval_type, round_num],
    )

    inserted = 0
    conn.execute("BEGIN")
    try:
        for _, row in df.iterrows():
            sid = int(row["학번"])
            for _, prob in problems.iterrows():
                p_num = int(prob["문제번호"])
                score = row.get(f"{p_num}_점수")
                deduct = row.get(f"{p_num}_감점")
                conn.execute(
                    """INSERT OR REPLACE INTO grades
                       (학번, 평가유형, 회차, 문제번호, 점수, 감점이유)
                       VALUES (?, ?, ?, ?, ?, ?)""",
                    (
                        sid,
                        eval_type,
                        round_num,
                        p_num,
                        None if pd.isna(score) else float(score),
                        None if pd.isna(deduct) else str(deduct),
                    ),
                )
                inserted += 1
        conn.commit()
    except Exception as e:
        conn.rollback()
        print(f"[ERROR] 롤백됨: {e}")
        conn.close()
        return

    conn.close()
    print(f"[OK] 채점 데이터 {inserted}건 등록 ({eval_type} {round_num}회차)")


def cmd_export_lms(args):
    eval_type, round_num = args.type, args.round
    conn = get_conn()

    df = pd.read_sql(
        """
        SELECT
            st.학번, st.이름,
            COALESCE(sub.제출여부, 'X') AS 제출여부,
            COALESCE(sub.지각유형, '2형') AS 지각유형,
            COALESCE(SUM(g.점수), 0) AS 총점
        FROM students st
        LEFT JOIN submissions sub
            ON st.학번 = sub.학번 AND sub.평가유형 = ? AND sub.회차 = ?
        LEFT JOIN grades g
            ON st.학번 = g.학번 AND g.평가유형 = ? AND g.회차 = ?
        GROUP BY st.학번, st.이름
        ORDER BY st.학번
        """,
        conn,
        params=[eval_type, round_num, eval_type, round_num],
    )

    df["최종점수"] = df["총점"].copy()
    mask_late = df["지각유형"] == "1형"
    mask_miss = (df["지각유형"] == "2형") | (df["제출여부"] == "X")
    df.loc[mask_late, "최종점수"] = (df.loc[mask_late, "총점"] * 0.9).round(1)
    df.loc[mask_miss, "최종점수"] = 0

    total_max = pd.read_sql(
        "SELECT SUM(만점) AS v FROM problems WHERE 평가유형=? AND 회차=?",
        conn,
        params=[eval_type, round_num],
    ).iloc[0]["v"]

    wb = Workbook()
    ws = wb.active
    ws.title = "LMS점수"

    headers = ["학번", "이름", "총점", "지각유형", "최종점수"]
    for c, h in enumerate(headers, 1):
        _styled_cell(ws, 1, c, h, fill=HEADER_FILL, font=HEADER_FONT)

    for r, (_, row) in enumerate(df.iterrows(), 2):
        _styled_cell(ws, r, 1, int(row["학번"]))
        _styled_cell(ws, r, 2, row["이름"])
        _styled_cell(ws, r, 3, row["총점"])
        _styled_cell(ws, r, 4, row["지각유형"])
        _styled_cell(ws, r, 5, row["최종점수"])

    for c, w in enumerate([14, 10, 8, 10, 10], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"

    out_path = Path(f"rounds/{eval_type}_{round_num}/lms_{eval_type}_{round_num}.xlsx")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    conn.close()

    avg = df["최종점수"].mean()
    print(f"[OK] LMS 점수 시트: {out_path}")
    print(f"     총만점: {total_max} | 평균: {avg:.1f} | 학생: {len(df)}명")


def cmd_export_feedback(args):
    eval_type, round_num = args.type, args.round
    conn = get_conn()

    # 감점이유가 있는 항목만 추출 (+ 미제출자 포함)
    df = pd.read_sql(
        """
        SELECT
            st.학번, st.이름,
            g.문제번호,
            p.만점,
            g.점수,
            g.감점이유,
            COALESCE(sub.제출여부, 'X') AS 제출여부,
            COALESCE(sub.지각유형, '2형') AS 지각유형
        FROM students st
        LEFT JOIN submissions sub
            ON st.학번 = sub.학번 AND sub.평가유형 = ? AND sub.회차 = ?
        LEFT JOIN grades g
            ON st.학번 = g.학번 AND g.평가유형 = ? AND g.회차 = ?
        LEFT JOIN problems p
            ON g.평가유형 = p.평가유형 AND g.회차 = p.회차
            AND g.문제번호 = p.문제번호
        ORDER BY st.학번, g.문제번호
        """,
        conn,
        params=[eval_type, round_num, eval_type, round_num],
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "피드백"

    headers = ["학번", "이름", "문제번호", "만점", "점수", "감점이유", "제출여부", "지각유형"]
    for c, h in enumerate(headers, 1):
        _styled_cell(ws, 1, c, h, fill=HEADER_FILL, font=HEADER_FONT)

    for r, (_, row) in enumerate(df.iterrows(), 2):
        _styled_cell(ws, r, 1, int(row["학번"]) if pd.notna(row["학번"]) else "")
        _styled_cell(ws, r, 2, row["이름"] if pd.notna(row["이름"]) else "")
        _styled_cell(
            ws, r, 3, int(row["문제번호"]) if pd.notna(row["문제번호"]) else ""
        )
        _styled_cell(ws, r, 4, int(row["만점"]) if pd.notna(row["만점"]) else "")
        _styled_cell(ws, r, 5, row["점수"] if pd.notna(row["점수"]) else 0)

        deduct_val = row["감점이유"] if pd.notna(row["감점이유"]) else ""
        fill = DEDUCT_FILL if deduct_val else None
        _styled_cell(ws, r, 6, deduct_val, fill=fill, align=WRAP)

        _styled_cell(ws, r, 7, row["제출여부"] if pd.notna(row["제출여부"]) else "")
        _styled_cell(ws, r, 8, row["지각유형"] if pd.notna(row["지각유형"]) else "")

    for c, w in enumerate([14, 10, 10, 8, 8, 25, 10, 10], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"

    out_path = Path(f"rounds/{eval_type}_{round_num}/feedback_{eval_type}_{round_num}.xlsx")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    conn.close()
    print(f"[OK] 피드백 시트: {out_path}")


def cmd_create_input(args):
    out_path = Path(args.file)
    if out_path.exists() and not args.force:
        print(f"[ERROR] {out_path} 이(가) 이미 존재합니다. 덮어쓰려면 --force 옵션을 사용하세요.")
        return

    wb = Workbook()

    # Sheet 1: 수강생
    ws1 = wb.active
    ws1.title = "수강생"
    headers1 = ["학번", "이름"]
    for c, h in enumerate(headers1, 1):
        _styled_cell(ws1, 1, c, h, fill=HEADER_FILL, font=HEADER_FONT)
    ws1.column_dimensions["A"].width = 14
    ws1.column_dimensions["B"].width = 12

    # Sheet 2: 문제정보
    ws2 = wb.create_sheet("문제정보")
    headers2 = ["평가유형", "회차", "문제번호", "만점", "정답_소수", "정답_분수", "채점기준"]
    for c, h in enumerate(headers2, 1):
        _styled_cell(ws2, 1, c, h, fill=HEADER_FILL, font=HEADER_FONT)
    for c, w in enumerate([10, 6, 10, 6, 10, 10, 25], 1):
        ws2.column_dimensions[get_column_letter(c)].width = w

    # Sheet 3: 제출현황
    ws3 = wb.create_sheet("제출현황")
    headers3 = ["학번", "평가유형", "회차", "제출여부", "제출시각", "지각유형"]
    for c, h in enumerate(headers3, 1):
        _styled_cell(ws3, 1, c, h, fill=HEADER_FILL, font=HEADER_FONT)
    for c, w in enumerate([14, 10, 6, 10, 20, 10], 1):
        ws3.column_dimensions[get_column_letter(c)].width = w

    wb.save(out_path)
    print(f"[OK] 입력 파일 생성: {out_path}")
    print(f"     시트: 수강생 / 문제정보 / 제출현황")


def cmd_stats(args):
    eval_type, round_num = args.type, args.round
    conn = get_conn()

    problems = pd.read_sql(
        "SELECT 문제번호, 만점 FROM problems WHERE 평가유형=? AND 회차=? ORDER BY 문제번호",
        conn,
        params=[eval_type, round_num],
    )

    print(f"\n{'='*50}")
    print(f"  {eval_type} {round_num}회차 — 문제별 통계")
    print(f"{'='*50}")

    for _, p in problems.iterrows():
        p_num = int(p["문제번호"])
        max_s = int(p["만점"])
        scores = pd.read_sql(
            "SELECT 점수 FROM grades WHERE 평가유형=? AND 회차=? AND 문제번호=? AND 점수 IS NOT NULL",
            conn,
            params=[eval_type, round_num, p_num],
        )
        if not scores.empty:
            avg = scores["점수"].mean()
            rate = avg / max_s * 100
            print(
                f"  문제{p_num:>2} | 만점: {max_s:>3}"
                f" | 평균: {avg:>5.1f} | 득점률: {rate:>5.1f}%"
            )
        else:
            print(f"  문제{p_num:>2} | 만점: {max_s:>3} | 데이터 없음")

    total_max = problems["만점"].sum()
    total = pd.read_sql(
        """
        SELECT st.학번, COALESCE(SUM(g.점수), 0) AS 총점
        FROM students st
        LEFT JOIN grades g
            ON st.학번 = g.학번 AND g.평가유형 = ? AND g.회차 = ?
        GROUP BY st.학번
        """,
        conn,
        params=[eval_type, round_num],
    )
    avg_total = total["총점"].mean()
    print(
        f"\n  전체 | 총만점: {total_max}"
        f" | 평균: {avg_total:.1f} | 득점률: {avg_total / total_max * 100:.1f}%"
    )
    print(f"{'='*50}\n")
    conn.close()


# ── CLI ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="확률통계 채점 시스템")
    sub = parser.add_subparsers(dest="cmd", required=True)

    p = sub.add_parser("create-input", help="빈 grades.xlsx 생성 (수강생/문제정보/제출현황 시트)")
    p.add_argument("--file", default=str(INPUT_EXCEL))
    p.add_argument("--force", action="store_true", help="기존 파일 덮어쓰기")

    sub.add_parser("init", help="DB 초기화 (최초 1회)")

    p = sub.add_parser("import-students", help="수강생 명단 등록")
    p.add_argument("--file", default=str(INPUT_EXCEL))

    p = sub.add_parser("import-problems", help="문제정보 등록")
    p.add_argument("--file", default=str(INPUT_EXCEL))

    p = sub.add_parser("import-submissions", help="제출현황 등록")
    p.add_argument("--file", default=str(INPUT_EXCEL))

    p = sub.add_parser("gen-template", help="채점 템플릿 생성")
    p.add_argument("--type", required=True, help="과제 / 퀴즈 / 시험")
    p.add_argument("--round", type=int, required=True, help="회차")

    p = sub.add_parser("import-grades", help="채점 결과 등록 (템플릿 → DB)")
    p.add_argument("--file", required=True, help="채점 완료 템플릿")
    p.add_argument("--type", default=None, help="평가유형 (선택)")
    p.add_argument("--round", type=int, default=None, help="회차 (선택)")

    p = sub.add_parser("export-lms", help="LMS 점수 시트 생성")
    p.add_argument("--type", required=True)
    p.add_argument("--round", type=int, required=True)

    p = sub.add_parser("export-feedback", help="학생별 피드백 시트 생성")
    p.add_argument("--type", required=True)
    p.add_argument("--round", type=int, required=True)

    p = sub.add_parser("stats", help="문제별 통계 출력 (터미널)")
    p.add_argument("--type", required=True)
    p.add_argument("--round", type=int, required=True)

    args = parser.parse_args()

    dispatch = {
        "create-input": cmd_create_input,
        "init": cmd_init,
        "import-students": cmd_import_students,
        "import-problems": cmd_import_problems,
        "import-submissions": cmd_import_submissions,
        "gen-template": cmd_gen_template,
        "import-grades": cmd_import_grades,
        "export-lms": cmd_export_lms,
        "export-feedback": cmd_export_feedback,
        "stats": cmd_stats,
    }
    dispatch[args.cmd](args)


if __name__ == "__main__":
    main()
